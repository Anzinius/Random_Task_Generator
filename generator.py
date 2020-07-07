import sys
import os
import random
import io
import csv
from PyQt5 import QtWidgets
from PyQt5 import uic
from PyQt5 import Qt
from PyQt5 import QtCore
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from datetime import datetime

# path to ui file
ui_path = os.path.dirname(os.path.abspath(__file__))
form_class = uic.loadUiType(os.path.join(ui_path, "main.ui"))[0]

# Make window not to shut down when exception occurs
def my_exception_hook(exctype, value, traceback):
    # Print the error and traceback
    print(exctype, value, traceback)
    # Call the normal Exception hook after
    sys._excepthook(exctype, value, traceback)
    # sys.exit(1)
# Back up the reference to the exceptionhook
sys._excepthook = sys.excepthook
# Set the exception hook to our wrapping function
sys.excepthook = my_exception_hook

class MyWindow(QtWidgets.QMainWindow, form_class):
    def __init__(self):
        super().__init__(),
        self.setupUi(self)

        self.btn_open_xlsx.clicked.connect(self.btnUploadClicked)
        self.btn_search.clicked.connect(self.btnGenerateClicked)
        self.btn_clear.clicked.connect(self.btnClearClicked)
        self.btn_download.clicked.connect(self.btnDownloadClicked)
        self.btn_print.clicked.connect(self.printPreview)
        
        self.xlsx = str()
        self.quarter = int()
        self.keyword = []
        self.member = []
        self.work = []
        self.result = []
        self.task = []

    def quarterisChecked(self):
        if self.input_qut_first.isChecked(): self.quarter = 1
        elif self.input_qut_second.isChecked(): self.quarter = 2
        elif self.input_qut_third.isChecked(): self.quarter = 3
        elif self.input_qut_fourth.isChecked(): self.quarter = 4
        else: self.quarter = 0

    def filterTaskByDate(self, file):
        sheetWork = file['work']
        for row in range(2, sheetWork.max_row + 1):
            if sheetWork.cell(row, self.quarter + 3).value == True:
                self.work.append([sheetWork.cell(row, 2).value, sheetWork.cell(row, 3).value])

    def setTaskByKeyword(self, file):
        sheetKeyword = file['keyword']
        keywords = []
        works = []
        for row in range(1,sheetKeyword.max_row+1):
            keywords.append(sheetKeyword.cell(row, 1).value)
        for i in range(len(self.work)):
            works.append(self.work[i][0])
        for a in keywords:
            A = []
            A.append(a)
            for b in works:
                self.task.append([A[0]+" "+b, b])

    def setTaskByPosition(self, file):
        sheetMember = file['member']
        sheetPosition = file['position']
        members = []
        positions = []
        for row in range(2, sheetMember.max_row + 1):
            members.append([sheetMember.cell(row, 1).value, sheetMember.cell(row, 2).value])
        for row in range(2, sheetPosition.max_row + 1):
            positions.append([sheetPosition.cell(row, 1).value, sheetPosition.cell(row, 2).value])
        for a, b in members:
            for x, y in positions:
                if b == y:
                    self.member.append([a, b, x])

        cnt = 0
        for a, b, c in self.member:
            for x, y in self.task:
                for m, n in self.work:
                    cnt += 1
                    print(cnt)
                    if m == y and int(n) == 0:
                        self.result.append([x, y, a, b, c])
                    elif m == y and int(n) > 0:
                        if c >= n:
                            self.result.append([x, y, a, b, c])
                    elif m == y and int(n) < 0:
                        if -c <= n :
                            self.result.append([x, y, a, b, c])
    
    def keyPressEvent(self, ev):
        if (ev.key() == QtCore.Qt.Key_C) and (ev.modifiers() & QtCore.Qt.ControlModifier): 
            self.copySelection()

    def printPreview(self):
        self.tbl_preview.clearContents()

        for row in range(len(self.member)):
            mem = QtWidgets.QTableWidgetItem() #
            mem.setText(self.member[row][0])
            mem.setTextAlignment(QtCore.Qt.AlignCenter)
            self.tbl_preview.setItem(row,0,mem)
            
            randTask = []
            for i in range(len(self.result)):
                if self.result[i][2] == self.member[row][0]:
                    randTask.append(self.result[i][0])
            task = QtWidgets.QTableWidgetItem()
            task.setText(random.choice(randTask))
            self.tbl_preview.setItem(row,1,task)
        self.tbl_preview.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers) #https://m.blog.naver.com/PostView.nhn?blogId=thenaru2&logNo=220788804430&proxyReferer=https:%2F%2Fwww.google.com%2F

    def copySelection(self):
        selection = self.tbl_preview.selectedIndexes() #승민 호진 정수
        if selection:
            rows = sorted(index.row() for index in selection)
            columns = sorted(index.column() for index in selection)
            rowcnt = rows[-1] - rows[0] + 1
            colcnt = columns[-1] - columns[0] + 1
            table = [[''] * colcnt for _ in range(rowcnt)]
            for index in selection:
                row = index.row() - rows[0]
                column = index.column() - columns[0]
                table[row][column] = index.data()
                stream = io.StringIO()
                csv.writer(stream).writerows(table)
                QtWidgets.QApplication.clipboard().setText(stream.getvalue())

    def btnClearClicked(self):
        self.input_upload_xlsx.setText('*.xlsx 업로드')
        self.label_result_value.setText('')
        self.tbl_preview.clearContents()
        self.quarter = 0
        #self.groupBox

    def btnGenerateClicked(self):
        # initialize
        self.keyword = []
        self.member = []
        self.work = []
        self.result = []
        self.task = []
        self.tbl_preview.clearContents()

        # error1
        if len(self.xlsx) == 0 :
            QtWidgets.QMessageBox.warning(self, "주의", "파일을 먼저 선택해주세요.    ")
            return
        # error2
        self.quarterisChecked()
        if self.quarter == 0 :
            QtWidgets.QMessageBox.warning(self, "주의", "파일을 먼저 선택해주세요.    ")
            return

        # read and generate
        load_file = load_workbook(self.xlsx, data_only=True)
        try:
            self.filterTaskByDate(load_file)
            self.setTaskByKeyword(load_file)
            self.setTaskByPosition(load_file)
            self.label_result_value.setText(str(len(self.result)))
            print("Done!")
        except OSError:
            QtWidgets.QMessageBox.critical(self, "경고", "            잘못된 파일입니다.    \n 파일의 확장자나 내용을 확인해주세요.      ")
        except TypeError:
            QtWidgets.QMessageBox.critical(self, "경고", "            잘못된 파일입니다.    \n 파일의 확장자나 내용을 확인해주세요.      ")

    def btnDownloadClicked(self):
        result_file = Workbook()
        sheet1 = result_file.active
        sheet1.title = 'result'
        idx = ["업무(결과값)", "수행 내용", "이름", "직급", "직급 NO."]
        for i in range(5):
            sheet1.cell(1, i+1).value = idx[i]
        for cell in self.result:
            sheet1.append(cell)
        result_file.save('./result' + datetime.today().strftime("%Y%m%d%H%M%S") + '.xlsx') 
        QtWidgets.QMessageBox.about(self, "저장완료", "파일이 저장되었습니다.")  

    def btnUploadClicked(self):
        fname = QtWidgets.QFileDialog.getOpenFileName(self)
        self.input_upload_xlsx.setText(fname[0])
        self.xlsx = fname[0]

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()